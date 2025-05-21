import pandas as pd
from datetime import datetime
import PySimpleGUI as sg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from fpdf import FPDF

# --- GUI for file selection ---
sg.theme("LightBlue2")
layout = [
    [sg.Text("Select Excel file:")],
    [sg.Input(key="-FILE-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
    [sg.Button("Process"), sg.Button("Exit")]
]

window = sg.Window("Excel Sales Report Generator", layout)

while True:
    event, values = window.read()
    if event == sg.WINDOW_CLOSED or event == "Exit":
        break
    if event == "Process":
        filename = values["-FILE-"]
        if filename == "":
            sg.popup_error("Please select an Excel file.")
            continue

        try:
            # Read Excel file into DataFrame
            df = pd.read_excel(filename)

            # Calculate total and average sales
            total_sales = df["Sales"].sum()
            average_sales = round(df["Sales"].mean(), 2)

            # Add Total and Average rows using concat to avoid column mismatch
            summary_rows = pd.DataFrame([
                {"Product": "Total", "Sales": total_sales},
                {"Product": "Average", "Sales": average_sales}
            ])
            df = pd.concat([df, summary_rows], ignore_index=True)

            # Create timestamped output filenames
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
            output_excel = f"cleaned_{timestamp}.xlsx"
            output_pdf = f"report_{timestamp}.pdf"

            # --- Create styled Excel report ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"

            # Add DataFrame to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(row)

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            align_center = Alignment(horizontal="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border

            # Style Total and Average rows
            for row_idx in [ws.max_row - 1, ws.max_row]:
                for cell in ws[row_idx]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    cell.border = thin_border

            # Style all data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = align_center
                    cell.border = thin_border

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(output_excel)

            # --- Create PDF report ---
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "Sales Report", ln=True, align='C')
            pdf.ln(10)

            pdf.set_font("Arial", 'B', 12)
            pdf.cell(90, 10, "Product", border=1, align="C")
            pdf.cell(50, 10, "Sales", border=1, align="C")
            pdf.ln()

            pdf.set_font("Arial", '', 12)
            for _, row in df.iterrows():
                pdf.cell(90, 10, str(row["Product"]), border=1, align="C")
                pdf.cell(50, 10, str(row["Sales"]), border=1, align="C")
                pdf.ln()

            pdf.output(output_pdf)

            sg.popup(f"✅ Excel report saved as {output_excel}\n✅ PDF report saved as {output_pdf}")

        except FileNotFoundError:
            sg.popup_error(f"File '{filename}' not found.")
        except KeyError as e:
            sg.popup_error(f"Missing column in Excel file: {e}")
        except Exception as e:
            sg.popup_error(f"Error: {e}")

window.close()
